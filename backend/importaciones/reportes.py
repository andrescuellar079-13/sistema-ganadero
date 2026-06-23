"""Generación del reporte de errores en XLSX.

Columnas (según spec): hoja, fila, nro_arete, campo, valor_recibido,
codigo_error, descripcion, recomendacion.
"""
import io

# Recomendaciones legibles por código de error.
RECOMENDACIONES = {
    "REQUERIDO": "Completá este campo obligatorio.",
    "FECHA_INVALIDA": "Usá el formato DD/MM/AAAA o AAAA-MM-DD.",
    "DECIMAL_INVALIDO": "Usá un número válido (ej: 380.5 o 380,5).",
    "ENTERO_INVALIDO": "Usá un número entero válido.",
    "VALOR_NEGATIVO": "El valor no puede ser negativo.",
    "CHOICE_INVALIDO": "Elegí uno de los valores permitidos para esta columna.",
    "ARETE_DUPLICADO_ARCHIVO": "El nro_arete está repetido dentro del archivo.",
    "ARETE_DUPLICADO_BD": "Ya existe un animal con ese arete. Usá un modo de actualización.",
    "ARETE_NO_EXISTE": "El animal no existe; agregalo en la hoja ANIMALES o revisá el arete.",
    "RAZA_NO_EXISTE": "Registrá la raza en el catálogo antes de importar.",
    "CATEGORIA_NO_EXISTE": "Registrá la categoría en el catálogo antes de importar.",
    "PARCELA_NO_EXISTE": "Creá la parcela en la hoja PARCELAS o revisá el nombre.",
    "PADRE_NO_EXISTE": "El padre debe existir en la finca o en la hoja ANIMALES.",
    "MADRE_NO_EXISTE": "La madre debe existir en la finca o en la hoja ANIMALES.",
    "PADRE_NO_MACHO": "El padre referenciado debe ser MACHO.",
    "MADRE_NO_HEMBRA": "La madre referenciada debe ser HEMBRA.",
    "HEADER_FALTANTE": "Falta una columna obligatoria. Descargá la plantilla.",
    "NO_EXISTE_ACTUALIZAR": "No existe un animal con ese arete para actualizar.",
    "FILA_DUPLICADA_BD": "El registro histórico ya existe; se omitió para no duplicar.",
}

COLUMNAS = [
    "hoja", "fila", "nro_arete", "campo", "valor_recibido",
    "codigo_error", "descripcion", "recomendacion",
]


def generar_reporte_errores(errores):
    """``errores``: lista de dicts con las claves de ``ImportacionGanaderaError``."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "ERRORES"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C62828")
    for idx, nombre in enumerate(COLUMNAS, start=1):
        celda = ws.cell(row=1, column=idx, value=nombre)
        celda.font = header_font
        celda.fill = header_fill
        ws.column_dimensions[celda.column_letter].width = 22
    ws.freeze_panes = "A2"

    for e in errores:
        codigo = e.get("codigo_error", "")
        ws.append([
            e.get("hoja", ""),
            e.get("numero_fila", ""),
            e.get("nro_arete", "") or "",
            e.get("campo", "") or "",
            e.get("valor", "") or "",
            codigo,
            e.get("mensaje", "") or "",
            RECOMENDACIONES.get(codigo, "Revisá el dato e intentá de nuevo."),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def recomendacion(codigo):
    return RECOMENDACIONES.get(codigo, "Revisá el dato e intentá de nuevo.")
