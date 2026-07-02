"""Utilidades compartidas por los tests de importación."""
import io
import tempfile

from openpyxl import Workbook

from importaciones import constantes


def construir_xlsx(hojas):
    """``hojas``: {"ANIMALES": [ {header: valor}, ... ], ...} → bytes XLSX."""
    wb = Workbook()
    wb.remove(wb.active)
    for nombre, filas in hojas.items():
        ws = wb.create_sheet(nombre)
        headers = constantes.headers(nombre)
        ws.append(headers)
        for fila in filas:
            ws.append([fila.get(h, "") for h in headers])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def construir_xlsx_crudo(hojas):
    """Igual que ``construir_xlsx`` pero con encabezados arbitrarios.

    ``hojas``: {"ANIMALES": {"headers": [...], "filas": [[c1, c2, ...], ...]}}.
    Sirve para probar el matching de alias/acentos.
    """
    wb = Workbook()
    wb.remove(wb.active)
    for nombre, contenido in hojas.items():
        ws = wb.create_sheet(nombre)
        ws.append(contenido["headers"])
        for fila in contenido["filas"]:
            ws.append(list(fila))
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def media_temporal():
    return tempfile.mkdtemp(prefix="imp_test_media_")
