"""Generación de la plantilla XLSX de importación.

La plantilla tiene una hoja por cada entrada de ``constantes.HOJAS`` (en el
orden de procesamiento) más una hoja "INSTRUCCIONES". La primera fila son los
encabezados; la segunda, una fila de ejemplo para guiar al usuario.
"""
import io

from . import constantes


def _estilos():
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E7D32")
    ejemplo_font = Font(italic=True, color="888888")
    wrap = Alignment(vertical="top", wrap_text=True)
    return header_font, header_fill, ejemplo_font, wrap


def _ancho(header):
    return max(14, min(40, len(header) + 6))


def generar_plantilla():
    """Devuelve los bytes de un XLSX con la plantilla de importación."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)  # quita la hoja por defecto

    header_font, header_fill, ejemplo_font, wrap = _estilos()

    # Hoja de instrucciones
    ws_info = wb.create_sheet("INSTRUCCIONES")
    ws_info.column_dimensions["A"].width = 100
    lineas = [
        "PLANTILLA DE IMPORTACIÓN — SISTEMA GANADERO",
        "",
        "• Completá las hojas que necesités. nro_arete es la referencia global en todas.",
        "• Fechas: DD/MM/AAAA o AAAA-MM-DD.  Decimales: punto o coma (380.5 o 380,5).",
        "• Sexo: MACHO | HEMBRA.  Estado: ACTIVO | VENDIDO | MUERTO | DESCARTE | MATADERO | BAJA.",
        "• Tipo producción: CARNE | LECHE | DOBLE_PROPOSITO.  Origen: NACIDO_FINCA | COMPRADO | DONADO.",
        "• Razas y categorías deben existir en el catálogo. Las parcelas se crean desde la hoja PARCELAS.",
        "• La finca se elige en el sistema, NO se incluye en el archivo.",
        "• La segunda fila de cada hoja es un EJEMPLO: podés borrarla o reemplazarla.",
        "",
    ]
    for col_data in constantes.ORDEN_HOJAS:
        ws_info.append([f"Hoja {col_data}: {constantes.HOJAS[col_data]['descripcion']}"])
    for i, linea in enumerate(lineas, start=1):
        ws_info.insert_rows(i)
        ws_info.cell(row=i, column=1, value=linea)

    # Hojas de datos
    for clave in constantes.ORDEN_HOJAS:
        ws = wb.create_sheet(clave)
        cols = constantes.columnas(clave)
        for idx, c in enumerate(cols, start=1):
            celda = ws.cell(row=1, column=idx, value=c["header"])
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = wrap
            ws.column_dimensions[
                ws.cell(row=1, column=idx).column_letter
            ].width = _ancho(c["header"])
            ejemplo = ws.cell(row=2, column=idx, value=c.get("ejemplo", ""))
            ejemplo.font = ejemplo_font
        ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


PLANTILLA_NOMBRE = "plantilla_importacion_ganadera.xlsx"
