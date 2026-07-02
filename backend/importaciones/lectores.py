"""Lectura de archivos XLSX/CSV a una estructura normalizada por hoja.

Salida de ``leer_archivo``::

    {
      "ANIMALES": {
        "headers": {"nro_arete", "sexo", ...},   # keys canónicas detectadas
        "filas": [
          {"numero_fila": 2, "valores": {"nro_arete": "AR-1", ...}},
          ...
        ],
      },
      ...
    }

Solo se devuelven las hojas reconocidas en ``constantes.HOJAS``. Las celdas se
entregan crudas (texto / número / fecha nativa); la normalización a tipos es
responsabilidad de ``validadores``.

Seguridad: el XLSX se abre con ``data_only=True`` → se leen los valores
cacheados, nunca se evalúan fórmulas.
"""
import csv
import io

from . import constantes


class ArchivoInvalidoError(Exception):
    """El archivo no se pudo abrir o no tiene un formato soportado."""


def _hoja_canonica(nombre):
    """Devuelve la clave de HOJAS que corresponde a un título de hoja, o None."""
    objetivo = (nombre or "").strip().upper()
    for clave in constantes.HOJAS:
        if clave.upper() == objetivo:
            return clave
    return None


def _fila_vacia(valores):
    return all(
        v is None or (isinstance(v, str) and v.strip() == "")
        for v in valores.values()
    )


def _mapear_fila(celdas, mapa_headers, indices):
    """Construye {key: valor} para una fila a partir de celdas posicionales."""
    valores = {}
    for idx, key in indices.items():
        valores[key] = celdas[idx] if idx < len(celdas) else None
    return valores


def _procesar_tabla(clave_hoja, encabezados, filas_datos, fila_offset):
    """Convierte encabezados + filas en la estructura normalizada de una hoja.

    Devuelve, además de ``headers``/``filas``, un ``mapeo`` con una entrada por
    cada columna del archivo: ``{"columna": <texto original>, "key": <key|None>}``
    para que el frontend muestre qué reconoció (y qué no) el importador.
    """
    mapa = constantes.header_a_key(clave_hoja)
    indices = {}
    headers_encontrados = set()
    mapeo = []
    for idx, encabezado in enumerate(encabezados):
        if encabezado is None or str(encabezado).strip() == "":
            continue
        key = mapa.get(constantes.normalizar_encabezado(encabezado))
        mapeo.append({"columna": str(encabezado).strip(), "key": key})
        # La primera columna que mapea a una key gana; las repetidas se ignoran.
        if key and key not in headers_encontrados:
            indices[idx] = key
            headers_encontrados.add(key)

    filas = []
    for offset, celdas in enumerate(filas_datos):
        valores = _mapear_fila(list(celdas), mapa, indices)
        if _fila_vacia(valores):
            continue
        filas.append({"numero_fila": fila_offset + offset, "valores": valores})

    return {"headers": headers_encontrados, "filas": filas, "mapeo": mapeo}


def _leer_xlsx(contenido):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependencia obligatoria
        raise ArchivoInvalidoError(
            "Falta la dependencia 'openpyxl' en el backend."
        ) from exc

    try:
        wb = load_workbook(
            io.BytesIO(contenido), data_only=True, read_only=True
        )
    except Exception as exc:
        raise ArchivoInvalidoError(
            "No se pudo leer el archivo Excel (¿está dañado?)."
        ) from exc

    resultado = {}
    for ws in wb.worksheets:
        clave = _hoja_canonica(ws.title)
        if not clave:
            continue
        filas_iter = ws.iter_rows(values_only=True)
        try:
            encabezados = list(next(filas_iter))
        except StopIteration:
            continue
        filas_datos = list(filas_iter)
        # En Excel la fila 1 son encabezados; los datos empiezan en la fila 2.
        resultado[clave] = _procesar_tabla(clave, encabezados, filas_datos, 2)
    wb.close()
    return resultado


def _leer_csv(contenido, nombre_hoja=None):
    texto = contenido.decode("utf-8-sig", errors="replace")
    muestra = texto[:4096]
    try:
        dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t")
    except csv.Error:
        dialecto = csv.excel
    lector = csv.reader(io.StringIO(texto), dialecto)
    filas = list(lector)
    if not filas:
        raise ArchivoInvalidoError("El archivo CSV está vacío.")

    encabezados = filas[0]
    # Un CSV solo transporta una hoja: se infiere cuál por sus encabezados.
    clave = nombre_hoja or _inferir_hoja(encabezados)
    if not clave:
        raise ArchivoInvalidoError(
            "No se reconocieron los encabezados del CSV. Descargá la plantilla."
        )
    return {clave: _procesar_tabla(clave, encabezados, filas[1:], 2)}


def _inferir_hoja(encabezados):
    nombres = {constantes.normalizar_encabezado(h) for h in encabezados if h}
    mejor, mejor_puntaje = None, 0
    for clave in constantes.HOJAS:
        mapa = set(constantes.header_a_key(clave).keys())
        puntaje = len(nombres & mapa)
        if puntaje > mejor_puntaje:
            mejor, mejor_puntaje = clave, puntaje
    return mejor if mejor_puntaje > 0 else None


def leer_archivo(contenido, nombre_archivo):
    """Lee bytes de un XLSX o CSV y devuelve la estructura por hoja."""
    nombre = (nombre_archivo or "").lower()
    if nombre.endswith(".csv"):
        return _leer_csv(contenido)
    if nombre.endswith((".xlsx", ".xlsm")):
        return _leer_xlsx(contenido)
    raise ArchivoInvalidoError(
        "Formato no soportado. Subí un archivo .xlsx o .csv."
    )
